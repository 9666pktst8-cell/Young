#!/usr/bin/env python3
# PI格式分析脚本

import pandas as pd
import openpyxl
import json
from openpyxl import load_workbook
import sys

def analyze_pi_excel(file_path):
    """分析PI Excel文件的结构和内容"""
    
    print("=" * 80)
    print("开始分析PI文件:", file_path)
    print("=" * 80)
    
    # 加载工作簿
    wb = load_workbook(file_path, data_only=True)
    
    # 获取所有工作表名称
    sheet_names = wb.sheetnames
    print(f"工作表数量: {len(sheet_names)}")
    print(f"工作表名称: {sheet_names}")
    
    # 分析第一个工作表（通常是主要PI内容）
    ws = wb[sheet_names[0]]
    
    # 收集关键信息
    pi_info = {
        "文件信息": {
            "文件名": file_path.split('/')[-1],
            "工作表数量": len(sheet_names),
            "工作表名称": sheet_names
        },
        "PI基本信息": {},
        "公司信息": {},
        "客户信息": {},
        "产品明细": [],
        "条款信息": {},
        "格式特点": []
    }
    
    # 读取单元格内容并分析
    print("\n" + "=" * 80)
    print("PI关键信息提取:")
    print("=" * 80)
    
    # 收集所有非空单元格的内容（前50行，前20列）
    data = []
    for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=20):
        row_data = []
        for cell in row:
            if cell.value:
                row_data.append(f"{cell.coordinate}: {cell.value}")
        if row_data:
            data.append(row_data)
    
    # 输出前20行数据供分析
    print("\n前20行数据预览:")
    for i, row in enumerate(data[:20], 1):
        print(f"行{i}: {row}")
    
    # 尝试识别关键字段
    keywords = {
        "PI编号": ["PI NO", "PI No", "PI No.", "PI Number", "PROFORMA INVOICE NO"],
        "日期": ["DATE", "Date", "Issued Date", "Date:"],
        "卖方": ["SELLER", "Seller", "Beneficiary", "Exporter"],
        "买方": ["BUYER", "Buyer", "Applicant", "Importer"],
        "产品描述": ["DESCRIPTION OF GOODS", "Product Description", "Description", "Item"],
        "数量": ["QUANTITY", "Quantity", "QTY", "Qty"],
        "单价": ["UNIT PRICE", "Unit Price", "Price", "USD"],
        "总价": ["AMOUNT", "Amount", "Total", "TOTAL"],
        "条款": ["TERMS", "Terms", "Payment Terms", "Delivery"],
        "银行信息": ["BANK", "Bank", "Bank Details", "Account"]
    }
    
    print("\n" + "=" * 80)
    print("识别到的关键字段:")
    print("=" * 80)
    
    found_keywords = {}
    for key, variants in keywords.items():
        found = []
        for row in data:
            for cell_info in row:
                for variant in variants:
                    if variant.lower() in cell_info.lower():
                        found.append(cell_info)
        if found:
            found_keywords[key] = found[:3]  # 只取前3个匹配
            print(f"{key}: {found[:3]}")
    
    # 分析表格结构
    print("\n" + "=" * 80)
    print("表格结构分析:")
    print("=" * 80)
    
    # 查找可能的表格区域（连续非空行）
    table_start = None
    table_end = None
    for i, row in enumerate(data):
        if len(row) >= 3:  # 假设表格行至少有3个单元格
            if table_start is None:
                table_start = i
            table_end = i
        elif table_start is not None:
            break
    
    if table_start is not None:
        print(f"疑似产品表格区域: 行{table_start+1} 到 行{table_end+1}")
    
    # 保存分析结果
    pi_info["关键词匹配"] = found_keywords
    pi_info["数据预览"] = data[:30]
    
    # 生成分析报告
    report_path = file_path.replace('.xlsx', '_analysis.txt')
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("PI文件分析报告\n")
        f.write("=" * 80 + "\n\n")
        f.write(f"分析文件: {file_path}\n")
        f.write(f"分析时间: {pd.Timestamp.now()}\n\n")
        
        f.write("1. 文件结构\n")
        f.write("-" * 40 + "\n")
        f.write(f"工作表数量: {len(sheet_names)}\n")
        f.write(f"工作表名称: {sheet_names}\n\n")
        
        f.write("2. 识别到的关键信息\n")
        f.write("-" * 40 + "\n")
        for key, matches in found_keywords.items():
            f.write(f"{key}:\n")
            for match in matches:
                f.write(f"  - {match}\n")
            f.write("\n")
        
        f.write("3. 数据预览\n")
        f.write("-" * 40 + "\n")
        for i, row in enumerate(data[:30], 1):
            f.write(f"行{i}: {row}\n")
    
    print(f"\n分析报告已保存到: {report_path}")
    
    return pi_info

def extract_pi_template(file_path):
    """提取PI模板格式"""
    
    print("\n" + "=" * 80)
    print("提取PI模板格式:")
    print("=" * 80)
    
    # 这里可以添加更详细的格式提取逻辑
    # 比如识别表头、列宽、合并单元格等
    
    template = {
        "文件类型": "Excel PI",
        "建议结构": [
            "1. 公司抬头（Logo、公司名称、地址、联系方式）",
            "2. PI标题和编号",
            "3. 买卖双方信息",
            "4. 产品明细表格",
            "5. 金额汇总",
            "6. 条款和条件",
            "7. 银行信息",
            "8. 签名和盖章"
        ],
        "常见字段": [
            "PI No./编号",
            "Date/日期",
            "Seller/卖方",
            "Buyer/买方",
            "Description of Goods/产品描述",
            "Quantity/数量",
            "Unit Price/单价",
            "Amount/金额",
            "Total Amount/总金额",
            "Payment Terms/付款条款",
            "Delivery Time/交货期",
            "Bank Details/银行信息"
        ],
        "格式特点": [
            "中英文对照",
            "表格形式的产品明细",
            "金额使用USD",
            "包含公司银行信息",
            "有签名盖章位置"
        ]
    }
    
    return template

if __name__ == "__main__":
    file_path = "/root/.openclaw/workspace/pi_templates/PI-B3-26002_14th_March_2026---0a70a83e-f0a2-435d-aaab-1934b0476d31.xlsx"
    
    try:
        # 分析PI文件
        pi_info = analyze_pi_excel(file_path)
        
        # 提取模板格式
        template = extract_pi_template(file_path)
        
        # 保存模板信息
        template_path = "/root/.openclaw/workspace/pi_templates/pi_template_format.json"
        with open(template_path, 'w', encoding='utf-8') as f:
            json.dump({
                "pi_info": pi_info,
                "template_format": template,
                "source_file": file_path,
                "analysis_date": str(pd.Timestamp.now())
            }, f, ensure_ascii=False, indent=2)
        
        print(f"\nPI模板格式已保存到: {template_path}")
        
    except Exception as e:
        print(f"分析过程中出现错误: {e}")
        import traceback
        traceback.print_exc()